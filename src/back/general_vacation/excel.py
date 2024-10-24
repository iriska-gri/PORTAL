import pandas as pd
from pandas.io.excel import ExcelWriter
import openpyxl as oxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Side, Border, PatternFill
from openpyxl.worksheet.formula import ArrayFormula
from pandas.api.types import CategoricalDtype
import urllib.request
from datetime import date
from .models import *
from datetime import datetime
from django.db.models import Count

class roundExcel():
        def __init__(self, us):
            self.df = pd.DataFrame.from_records(us)
            self.wb = Workbook()
         
            pass

        def excround(self):
            
            self.df['tabel'] = self.df['login'].str.split('-', expand = True)[1]+self.df['login'].str.split('-', expand = True)[2]
            file_path = 'media/DailyRoundUp.xlsx'
            thins = Side(border_style="thin")
            ws = self.wb.active
            filters = ws.auto_filter
            filters.ref = "A6:N15"
            ws.title = "Список"
            ws['B2'] = 'Список сотрудников'
            ws['B3'] = 'МИ ФНС России по ЦОД № 2'
            ws['B4'] = 'по подразделениям'

            ws['B6'] = 'Таб.'
            ws['C6'] = 'Фамилия'
            ws['D6'] = 'Имя'
            ws['E6'] = 'Очество'

            for col in ws.iter_cols(min_row=6, max_row=6, min_col=2, max_col=5):
                for cell in col:
                    cell.font = Font(name='Times New Roman', size=11, bold=True)
                    cell.border = Border(top=thins, bottom=thins, left=thins, right=thins) 

            ws['G6'] = 'Отпуск'
            ws['I6'] = 'Болен'
            ws['K6'] = 'Удаленка'
            ws['M6'] = 'Командировка'


            

            heads =['G6', 'I6', 'K6', 'M6']
            for x in heads: 

                ws[x].fill = PatternFill(fgColor="00FFFF00", fill_type = "solid")
                ws[x].font = Font(name='Times New Roman', size=10, bold=True)
                ws[x].border = Border(top=thins, bottom=thins, left=thins, right=thins) 
                ws[x].alignment = Alignment(horizontal='center')

        
            ws['I1'] = 'На '+ date.today().strftime("%d.%m.%Y")
    
            numb = 1

            column_widths = [29, 53, 96, 96, 112, 28, 222, 29, 148, 28, 145, 33, 148, 28]
            for i, width in enumerate(column_widths):
                column_letter = oxl.utils.get_column_letter(i+1)
                ws.column_dimensions[column_letter].width = self.convert_width(width)

            

            for col in ws.iter_cols(min_row=2, max_row=5, min_col=2, max_col=5):
                for cell in col:
                    cell.font = Font(name='Times New Roman', size=12, bold=True)
                    cell.alignment = Alignment(horizontal='center')

            ws['I1'].font = Font(name='Times New Roman', size=14, bold=True)
            ws['I1'].alignment = Alignment(horizontal='left')

            
            for x in range(3):
                ws.merge_cells(start_row=2+x, start_column=2, end_row=2+x, end_column=6)

            
            k=7    
        
    
            for i in self.df['department'].unique():

                ws.cell(row = k, column = 2).value = i

                ws.cell(row = k, column = 2).font = Font(name='Times New Roman', size=10, bold=True)
                ws.merge_cells(start_row=k, start_column=2, end_row=k, end_column=12)

                l = pd.DataFrame(self.df[(self.df.department==i)][['tabel', 'familyname', 'name', 'parentname']])
               
           
                        
                zap = [7, 9, 11, 13]
                for x in  range(len(l)):
                    k=k+1
                
                    ws.cell(row = k, column = 1).value = numb
                    ws.cell(row = k, column = 1).font = Font(name='Times New Roman', size=10, bold=True)

                    ws.cell(row = k, column = 2).value = l.iloc[x]['tabel']
                    ws.cell(row = k, column = 3).value = l.iloc[x]['familyname']
                    ws.cell(row = k, column = 4).value = l.iloc[x]['name']
                    ws.cell(row = k, column = 5).value = l.iloc[x]['parentname']
                    ws.cell(row = k, column = 6).value = '=If(OR(H'+str(k)+'=1, J'+str(k)+'=1, L'+str(k)+'=1, N'+str(k)+'=1),"",1)'

                    for h in zap:
                        ws.cell(row = k, column = h).border = Border(top=thins, bottom=thins, left=thins, right=thins) 
                        ws.cell(row = k, column = h).font = Font(name='Times New Roman', size=10)
                        ws.cell(row = k, column = h).alignment = Alignment(horizontal='center')
                    
                    
                    for v in range(4):
                        ws.cell(row = k, column = 2+v).font = Font(name='Times New Roman', size=10)
                        ws.cell(row = k, column = 2+v).border = Border(top=thins, bottom=thins, left=thins, right=thins) 
                        
                
                    numb=numb+1
              
                k=k+1

            col=6
            ws.cell(row = k+1, column=col).value = len(self.df)
            ws.cell(row = k+1, column=col).fill = PatternFill(fgColor="00FFFF00", fill_type = "solid")
            ws.cell(row = k+1, column=col).font = Font(name='Times New Roman', size=13, bold=True)
            
            col+=1
            ws.cell(row = k+1, column=col).value = "ВСЕГО отпуск="
            ws.cell(row = k+1, column=col).fill = PatternFill(fgColor="00FFFF00", fill_type = "solid")
            ws.cell(row = k+1, column=col).font = Font(name='Times New Roman', size=10, bold=True)
            
            col+=1
            ws.cell(row = k+1, column=col).value = "=SUM(H6:H"+str(k-1)+")"
            ws.cell(row = k+1, column=col).fill = PatternFill(fgColor="00FFFF00", fill_type = "solid")
            ws.cell(row = k+1, column=col).font = Font(name='Times New Roman', size=10, bold=True)
            
            val = ["Всего больных=", "=SUM(J6:J"+str(k-1)+")", "Всего удаленка=", "=SUM(L6:L"+str(k-1)+")", "Всего в командировке=", "=SUM(N6:N"+str(k-1)+")"]
            for x in val:
                col+=1
                self.resultpos(ws, x, k, col)

            self.wb.save(file_path)
            return file_path
         
        def convert_width(self, width):
            """
            Конвертируем ширину столбца Excel в ширину столбца openpyxl
            """
            PIXELS_PER_CHAR = 7
            return (width) / PIXELS_PER_CHAR
        
        def resultpos(self, ws, value, k, col):
            # Строка 88 подведение итогов
            ws.cell(row = k+1, column=col).value = value
            ws.cell(row = k+1, column=col).fill = PatternFill(fgColor="8EA9DB", fill_type = "solid")
            ws.cell(row = k+1, column=col).font = Font(name='Times New Roman', size=10, bold=True)
        
        def excelvacation(self):
             
           pass


class loadExcel():
        def __init__(self, us, files):
            self.file = files
            self.dfuser = pd.DataFrame.from_records(us)

            

        def roundfilse(self):
                ed = pd.read_excel(self.file)
                df = pd.DataFrame(ed)
                datas = df.columns[8].split(" ")[1]
                
                df.columns = ["a", "b", "c", "d", "e", "f", "g", "h", "i", "j", "k", "l", "m", "n"]
                df=df.iloc[6:]
                df= df[df.b.notna()]
                df= df[df.a.notna()]
                df['b']=df['b'].str[-3:]
                df=df[['b', 'f', 'h', 'j','l', 'n']]

                df['b'] = df['b'].astype(int)

                types = ['f', 'h', 'j','l', 'n']
                for n in types:
                    df[n] = df[n].fillna(0) 
                    df[n] = df[n].astype(int)

                df3 = pd.merge(df, self.dfuser, left_on= 'b', right_on="id")
                df3=df3[['login', 'f', 'h', 'j','l','n']]

                for ind in df3.index:
        
                    keys = ''
                    if df3['h'][ind] == 1:
                        keys = 4 # В отпуске
                    elif df3['j'][ind] == 1:
                        keys = 5 # Болен
                    elif df3['f'][ind] == 1:
                        keys = 1 # На работе
                    elif df3['l'][ind] == 1:
                        keys = 2 # Удаленка
                    elif df3['n'][ind] == 1:
                        keys = 3 # Командировка    
                    else:
                        keys = 6
                    base=BasePosition.objects.get(pk=keys)
                    DailyRoundUp.objects.update_or_create(todaysDate=datetime.strptime(datas, '%d.%m.%Y').strftime('%Y-%m-%d'),
                                                        user=df3['login'][ind], defaults={'position':base})
                
                # pos =  DailyRoundUp.objects.values('position__name', 'position').filter(position=6,todaysDate=date.today()).annotate(Count('user'))
                
        
            
                return ''
        
        def vacationsExcel(self):
            ed = pd.read_excel(self.file, sheet_name='Запланированные отпуска')
            df = pd.DataFrame(ed)
            df.columns = ["a", "b", "c", "d", "e", "f", "g"]
            df=df[['b', 'c', 'd', 'e','f']]
            df3 = pd.merge(df, self.dfuser, left_on= 'b', right_on="fio")
            df3=df3[['login', 'c', 'd', 'e','f']]
            df3.loc[(df3['e'] == 'Осн'), 'e'] = 'os'
            df3.loc[(df3['e'] == 'Доп н/д'), 'e'] = 'nd'
            df3.loc[(df3['e'] == 'Доп в/л'), 'e'] = 'vl'
            df3['d'] = df3['d'].astype(int)
        
            for ind in df3.index:

                GeneralVacation.objects.create(login=df3['login'][ind], planned_date=df3['c'][ind], count_days=df3['d'][ind], type_of_vacation=df3['e'][ind], start_work_period=df3['f'][ind])
        